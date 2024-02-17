""" 3x input of Ichi Parameters ->
calculated and saved in new file in additional columns ->
build chart using calculated values to see Ichi """

import openpyxl as xl
from openpyxl.chart import AreaChart, StockChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.updown_bars import UpDownBars
from openpyxl.chart.data_source import NumData, NumVal

filename_with_format = str(input("Name Of File To Process: "))
file_format = str(input("File Format: "))
filename = filename_with_format + "." + file_format

par1 = int(input("Tenkan Value: "))
par2 = int(input("Kijun Value: "))
par3 = int(input("Lagging Value: "))

wb = xl.load_workbook(filename)
sheet = wb[wb.sheetnames[0]]


class Ichimoku:
    def tenkan(self):

        max_prices = []
        min_prices = []
        max_list = []
        min_list = []

        ts_value_final = []

        for row in range(2, sheet.max_row - (par1 - 2)):
            for values in range(row, row + par1):
                max_val = sheet.cell(values, 3).value
                min_val = sheet.cell(values, 4).value
                max_prices.append(max_val)
                min_prices.append(min_val)

            ts_max = max(max_prices)
            ts_min = min(min_prices)
            max_list.append(ts_max)
            min_list.append(ts_min)
            max_prices.clear()
            min_prices.clear()

        for items in range(0, len(max_list)):
            ts_val = (max_list[items] + min_list[items]) / 2
            ts_value_final.append(ts_val)

        ts_value = ts_value_final
        return ts_value

    def kijun(self):

        max_prices = []
        min_prices = []
        max_list = []
        min_list = []

        ks_value_final = []

        for row in range(2, sheet.max_row - (par2 - 2)):
            for values in range(row, row + par2):
                max_val = sheet.cell(values, 3).value
                min_val = sheet.cell(values, 4).value
                max_prices.append(max_val)
                min_prices.append(min_val)

            ks_max = max(max_prices)
            ks_min = min(min_prices)
            max_list.append(ks_max)
            min_list.append(ks_min)
            max_prices.clear()
            min_prices.clear()

        for items in range(0, len(max_list)):
            ks_val = (max_list[items] + min_list[items]) / 2
            ks_value_final.append(ks_val)

        ks_value = ks_value_final
        return ks_value

    def ssa(self):

        ssa = []
        ts_value = self.tenkan()
        ks_value = self.kijun()
        ssb_value = self.ssb()

        for items in range(0, len(ks_value)):
            ssa_val = (ts_value[items - par1 + par2] + ks_value[items]) / 2
            ssa.append(ssa_val)

        ssa_val = ssa
        return ssa_val

    def ssb(self):

        max_prices = []
        min_prices = []
        max_list = []
        min_list = []

        ssb_value_final = []

        for row in range(2, sheet.max_row - (par3 - 2)):
            for values in range(row, row + par3):
                max_val = sheet.cell(values, 3).value
                min_val = sheet.cell(values, 4).value
                max_prices.append(max_val)
                min_prices.append(min_val)

            ssb_max = max(max_prices)
            ssb_min = min(min_prices)
            max_list.append(ssb_max)
            min_list.append(ssb_min)
            max_prices.clear()
            min_prices.clear()

        for items in range(0, len(max_list)):
            ssb_val = (max_list[items] + min_list[items]) / 2
            ssb_value_final.append(ssb_val)

        ssb_value = ssb_value_final
        return ssb_value

    def chikou(self):

        chikou_value = []

        for prices in range(par2 + 1, sheet.max_row + 1):
            ch_val = sheet.cell(prices, 5).value
            chikou_value.append(ch_val)

        chikou = chikou_value
        return chikou

    def diff(self):

        diff_value = []
        ssa_value = self.ssa()
        ssb_value = self.ssb()

        for diffs in range(0, len(ssb_value)):
            dif = ssa_value[diffs] - ssb_value[diffs]
            diff_value.append(dif)

        diff_val = diff_value
        return diff_val

    def process_workbook(self):

        ts = self.tenkan()
        ks = self.kijun()
        ssa = self.ssa()
        ssb = self.ssb()
        ch = self.chikou()
        df = self.diff()

        for row in range(1 + par1, len(ts) + 1 + par1):
            ts_cell = sheet.cell(row, 9)
            ts_cell.value = ts[row - 1 - par1]

        for row in range(1 + par2, len(ks) + 1 + par2):
            ks_cell = sheet.cell(row, 10)
            ks_cell.value = ks[row - 1 - par2]

        for row in range(1 + par3, len(ks) + 1 + par3):
            ssa_cell = sheet.cell(row, 11)
            ssa_cell.value = ssa[row - 1 - par3]

        for row in range(1 + par2 + par3, len(ssb) + 1 + par2 + par3):
            ssb_cell = sheet.cell(row, 12)
            ssb_cell.value = ssb[row - 1 - par2 - par3]

        for row in range(2, len(ch) + 2):
            ch_cell = sheet.cell(row, 13)
            ch_cell.value = ch[row - 2]

        for row in range(1 + par3 + par2, len(ssb) + 1 + par3 + par2):
            diff_cell = sheet.cell(row, 14)
            diff_cell.value = df[row - 1 - par3 - par2]

        ts_cell_rename = sheet.cell(1, 9)
        ts_cell_rename.value = "Tenkan"

        ks_cell_rename = sheet.cell(1, 10)
        ks_cell_rename.value = "Kijun"

        ssa_cell_rename = sheet.cell(1, 11)
        ssa_cell_rename.value = "SSA"

        ssb_cell_rename = sheet.cell(1, 12)
        ssb_cell_rename.value = "SSB"

        ch_cell_rename = sheet.cell(1, 13)
        ch_cell_rename.value = "Chikou"

        diff_cell_rename = sheet.cell(1, 14)
        diff_cell_rename.value = "Difference"

        """ Pointing Data For Chart """

        values = Reference(
            sheet, min_row=1, max_row=sheet.max_row, min_col=9, max_col=13
        )
        candles = Reference(
            sheet, min_row=1, max_row=sheet.max_row, min_col=2, max_col=5
        )
        span = Reference(
            sheet, min_row=1, max_row=sheet.max_row, min_col=11, max_col=12
        )

        """ Creating Candles Chart """

        candle_chart = StockChart()
        candle_chart.add_data(candles, titles_from_data=True)

        for s in candle_chart.series:
            s.graphicalProperties.line.noFill = True

        candle_chart.hiLowLines = ChartLines()
        candle_chart.upDownBars = UpDownBars()

        pts = [NumVal(idx=i) for i in range(len(candles) - 1)]
        cache = NumData(pt=pts)
        candle_chart.series[-1].val.numRef.numCache = cache

        """ Creating Span Chart """

        span_chart = AreaChart()
        span_chart.add_data(span, titles_from_data=True)

        """ Creating Ichimoku Chart """

        chart = LineChart()  # Creating instance of object LineChart
        chart.title = "Ichimoku"
        chart.add_data(
            values, titles_from_data=True
        )  # Adding data to function LineChart
        chart.x_axis.title = "Time"
        chart.y_axis.title = "Price"

        """ Showing Each Line On Chart """

        s4 = chart.series[0]
        s4.graphicalProperties.line.solidFill = "FF0000"

        s5 = chart.series[1]
        s5.graphicalProperties.line.solidFill = "000080"

        s6 = chart.series[2]
        s6.graphicalProperties.line.solidFill = "00FF00"

        s7 = chart.series[3]
        s7.graphicalProperties.line.solidFill = "008000"

        s8 = chart.series[4]
        s8.graphicalProperties.line.solidFill = "660066"

        # for row in range(1 + par3 + par2, len(ssb) + 1 + par3 + par2):
        #     ssa_cell = sheet.cell(row, 11)
        #     ssa_cell.value = ssa[row - 1 - par2 - par2]
        #     ssb_cell = sheet.cell(row, 12)
        #     ssb_cell.value = ssb[row - 1 - par3 - par2]
        #     if ssa[row - 1 - par2 - par2] > ssb[row - 1 - par3 - par2]:
        #         span_chart.layout = Layout(ManualLayout(y=0.25))
        #     if ssa[row - 1 - par2 - par2] < ssb[row - 1 - par3 - par2]:
        #         span_chart.layout = Layout(ManualLayout(y=0.25))

        # ssa_fill = span_chart.series[0]
        # ssa_fill.graphicalProperties.solidFill = "99CC04"

        # ssb_fill = span_chart.series[1]
        # ssb_fill.graphicalProperties.solidFill = "000000"
        # ssb_fill.group = "stacked"

        # ch_fill = span_chart.series[2]
        # ch_fill.group = "stacked"

        # diff_fill = span_chart.series[3]
        # diff_fill.group = "stacked"
        # diff_fill.graphicalProperties.solidFill = "99CC04"

        chart += candle_chart
        # chart += span_chart

        sheet.add_chart(chart, "p2")  # Creating chart
        chart.height = 75
        chart.width = 150

        for ws in wb.worksheets:
            ws.sheet_view.zoomScale = 60

        wb.save("E:\\Code\\Java\\Java_Apps\\ichimoku\\data\\TSLA_Ichi_Update.xlsx")


ich = Ichimoku()

ich.process_workbook()

# print(f'Tenkan Value Is: {ich.tenkan()}')
# print(f'Kijun Value Is: {ich.kijun()}')
# print(len(ich.chikou()))
# print(f'SSA Value Is: {ich.ssa()}')
# print(f'SSB Value Is: {ich.ssb()}')
# print(f'Chikou Value Is: {ich.chikou()}')
